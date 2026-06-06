"""
sample_data_generator.py
Generates a realistic sample Excel workbook for testing the AI Financial Model.
Run this first to create demo_data.xlsx
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_sample_data(output_path: str = "demo_data.xlsx"):
    np.random.seed(42)

    # --- Revenue & P&L Sheet ---
    quarters = pd.date_range(start="2019-01-01", periods=24, freq="QS")
    base_revenue = 1_000_000
    growth_rates = np.random.normal(0.05, 0.02, 24).clip(0.01, 0.12)

    revenue = [base_revenue]
    for g in growth_rates[1:]:
        revenue.append(revenue[-1] * (1 + g))

    cogs_pct      = np.random.normal(0.42, 0.02, 24).clip(0.35, 0.55)
    opex_pct      = np.random.normal(0.28, 0.01, 24).clip(0.20, 0.36)
    tax_rate      = 0.21

    gross_profit  = [r * (1 - c) for r, c in zip(revenue, cogs_pct)]
    ebitda        = [gp - r * o for gp, r, o in zip(gross_profit, revenue, opex_pct)]
    da            = [r * 0.05 for r in revenue]
    ebit          = [e - d for e, d in zip(ebitda, da)]
    interest      = [r * 0.01 for r in revenue]
    ebt           = [e - i for e, i in zip(ebit, interest)]
    net_income    = [e * (1 - tax_rate) for e in ebt]

    pnl_df = pd.DataFrame({
        "Date":          quarters,
        "Revenue":       revenue,
        "COGS":          [r * c for r, c in zip(revenue, cogs_pct)],
        "Gross_Profit":  gross_profit,
        "EBITDA":        ebitda,
        "D&A":           da,
        "EBIT":          ebit,
        "Interest":      interest,
        "EBT":           ebt,
        "Net_Income":    net_income,
    })

    # --- Balance Sheet ---
    cash_ratio         = np.random.normal(0.15, 0.02, 24).clip(0.08, 0.25)
    receivables_ratio  = np.random.normal(0.12, 0.01, 24).clip(0.08, 0.18)
    inventory_ratio    = np.random.normal(0.10, 0.01, 24).clip(0.06, 0.16)
    ppe_ratio          = np.random.normal(0.30, 0.02, 24).clip(0.22, 0.40)
    debt_ratio         = np.random.normal(0.25, 0.02, 24).clip(0.15, 0.40)

    total_assets = [r * 2.5 for r in revenue]
    bs_df = pd.DataFrame({
        "Date":              quarters,
        "Cash":              [ta * c for ta, c in zip(total_assets, cash_ratio)],
        "Accounts_Receivable": [ta * r for ta, r in zip(total_assets, receivables_ratio)],
        "Inventory":         [ta * i for ta, i in zip(total_assets, inventory_ratio)],
        "PP&E":              [ta * p for ta, p in zip(total_assets, ppe_ratio)],
        "Total_Assets":      total_assets,
        "Short_Term_Debt":   [ta * d * 0.3 for ta, d in zip(total_assets, debt_ratio)],
        "Long_Term_Debt":    [ta * d * 0.7 for ta, d in zip(total_assets, debt_ratio)],
        "Total_Equity":      [ta * (1 - d) for ta, d in zip(total_assets, debt_ratio)],
    })

    # --- Cash Flow Sheet ---
    capex_ratio = np.random.normal(0.07, 0.01, 24).clip(0.04, 0.12)
    cf_df = pd.DataFrame({
        "Date":                  quarters,
        "Operating_Cash_Flow":   [ni + d + r * 0.02 for ni, d, r in zip(net_income, da, revenue)],
        "CapEx":                 [-r * c for r, c in zip(revenue, capex_ratio)],
        "Free_Cash_Flow":        [ni + d + r * 0.02 - r * c
                                  for ni, d, r, c in zip(net_income, da, revenue, capex_ratio)],
        "Dividends_Paid":        [-r * 0.015 for r in revenue],
    })

    # --- KPI Sheet ---
    kpi_df = pd.DataFrame({
        "Date":                 quarters,
        "Revenue_Growth_QoQ":  [0.0] + [(revenue[i] - revenue[i-1]) / revenue[i-1]
                                         for i in range(1, 24)],
        "Gross_Margin":        [gp / r for gp, r in zip(gross_profit, revenue)],
        "EBITDA_Margin":       [e / r for e, r in zip(ebitda, revenue)],
        "Net_Margin":          [ni / r for ni, r in zip(net_income, revenue)],
        "ROE":                 [ni / eq for ni, eq in
                                zip(net_income, bs_df["Total_Equity"])],
        "Debt_to_Equity":      [(s + l) / eq for s, l, eq in
                                zip(bs_df["Short_Term_Debt"], bs_df["Long_Term_Debt"],
                                    bs_df["Total_Equity"])],
    })

    # Write to Excel with formatting
    wb = Workbook()

    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="1F4E79")
    num_fill_even  = PatternFill("solid", start_color="EBF3FB")
    border_side    = Side(style="thin", color="B8CCE4")
    thin_border    = Border(bottom=border_side)
    center         = Alignment(horizontal="center")

    def write_sheet(ws, df, title):
        ws.title = title
        ws.column_dimensions["A"].width = 14

        # Title row
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " "))
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            fill = num_fill_even if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(value, pd.Timestamp):
                    cell.value        = value.strftime("%Y-%m-%d")
                    cell.alignment    = center
                elif isinstance(value, float):
                    cell.value        = round(value, 4)
                    cell.alignment    = Alignment(horizontal="right")
                else:
                    cell.value        = value
                if fill:
                    cell.fill = fill
                cell.border = thin_border

    sheets_data = [
        (pnl_df,  "P&L"),
        (bs_df,   "Balance_Sheet"),
        (cf_df,   "Cash_Flow"),
        (kpi_df,  "KPIs"),
    ]

    first = True
    for df, title in sheets_data:
        if first:
            ws    = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        write_sheet(ws, df, title)

    wb.save(output_path)
    print(f"✅  Sample data saved to: {output_path}")
    print(f"    Sheets: P&L, Balance_Sheet, Cash_Flow, KPIs")
    print(f"    Rows per sheet: {len(pnl_df)} quarters (2019–2024)")


if __name__ == "__main__":
    generate_sample_data()
