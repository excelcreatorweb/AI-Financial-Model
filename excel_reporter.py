"""
excel_reporter.py
Exports forecast results to a professional, color-coded Excel workbook.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from pathlib import Path

from forecaster import ForecastResult, ScenarioBundle, KPICalculator, EnsembleForecaster


# ------------------------------------------------------------------ #
#  Style constants (industry-standard coloring)                       #
# ------------------------------------------------------------------ #
C_HEADER_BG    = "1F4E79"   # dark blue header
C_HEADER_FG    = "FFFFFF"   # white text
C_INPUT_FG     = "0000FF"   # blue = hardcoded inputs
C_FORMULA_FG   = "000000"   # black = formulas
C_LINK_FG      = "008000"   # green = cross-sheet links
C_ACCENT       = "BDD7EE"   # light blue rows
C_BULL         = "C6EFCE"   # green fill - positive
C_BEAR         = "FFC7CE"   # red fill - negative
C_TITLE_BG     = "2E75B6"   # medium blue title row
C_SECTION      = "D9E1F2"   # section separator

THIN  = Side(style="thin",   color="B8CCE4")
THICK = Side(style="medium", color="2E75B6")


def _border(top=False, bottom=True, left=False, right=False):
    return Border(
        top    = THICK if top else THIN if top else Side(style=None),
        bottom = THICK if bottom else THIN if bottom else Side(style=None),
        left   = THICK if left else THIN if left else Side(style=None),
        right  = THICK if right else THIN if right else Side(style=None),
    )


def _hdr(cell, text, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.value      = text
    cell.font       = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill       = PatternFill("solid", start_color=bg)
    cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border     = Border(bottom=THIN, top=THIN, left=THIN, right=THIN)


def _num(cell, value, color=C_FORMULA_FG, fmt="#,##0;(#,##0);-"):
    cell.value      = value
    cell.font       = Font(color=color, size=10, name="Arial")
    cell.number_format = fmt
    cell.alignment  = Alignment(horizontal="right")


def _pct(cell, value, color=C_FORMULA_FG):
    _num(cell, value, color, fmt="0.0%;(0.0%);-")


def _col(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ------------------------------------------------------------------ #
#  Sheet builders                                                      #
# ------------------------------------------------------------------ #
def _summary_sheet(ws, series: pd.Series, results: dict,
                   bundle: ScenarioBundle, col_name: str):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"AI Financial Forecast Report — {col_name}"
    c.font      = Font(bold=True, size=16, color=C_HEADER_FG, name="Arial")
    c.fill      = PatternFill("solid", start_color=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # KPI boxes
    kpi = KPICalculator.full_report(series)
    best = min(results.values(), key=lambda r: r.mape) if results else None

    labels = ["Periods", "Start Value", "End Value", "Total Growth",
              "CAGR", "Mean", "Std Dev", "Best Model MAPE"]
    values = [
        kpi["periods"], kpi["start_value"], kpi["end_value"],
        kpi["total_growth_pct"] / 100, kpi["cagr_pct"] / 100,
        kpi["mean"], kpi["std_dev"],
        best.mape / 100 if best else 0,
    ]
    fmts = ["#,##0", "#,##0", "#,##0", "0.0%", "0.0%",
            "#,##0", "#,##0", "0.0%"]

    _hdr(ws["A3"], "Metric", bg=C_TITLE_BG)
    _hdr(ws["B3"], "Value",  bg=C_TITLE_BG)
    for i, (lbl, val, fmt) in enumerate(zip(labels, values, fmts), 4):
        ws[f"A{i}"].value = lbl
        ws[f"A{i}"].font  = Font(bold=True, size=10, color="2C3E50", name="Arial")
        ws[f"A{i}"].fill  = PatternFill("solid", start_color=(C_ACCENT if i % 2 == 0 else "FFFFFF"))
        _num(ws[f"B{i}"], val, fmt=fmt)
        ws[f"B{i}"].fill  = PatternFill("solid", start_color=(C_ACCENT if i % 2 == 0 else "FFFFFF"))

    # Model accuracy table
    _hdr(ws["D3"], "Model",         bg=C_TITLE_BG)
    _hdr(ws["E3"], "MAPE %",        bg=C_TITLE_BG)
    _hdr(ws["F3"], "RMSE",          bg=C_TITLE_BG)
    _hdr(ws["G3"], "Forecast CAGR", bg=C_TITLE_BG)

    for i, (name, result) in enumerate(results.items(), 4):
        fc_cagr = KPICalculator.compound_forecast_growth(result)
        ws[f"D{i}"].value = name
        ws[f"D{i}"].font  = Font(size=10, name="Arial")
        is_best = (best and name == best.model_name)
        fill = PatternFill("solid", start_color=(C_BULL if is_best else "FFFFFF"))
        for col in "DEFG":
            ws[f"{col}{i}"].fill = fill
        _num(ws[f"E{i}"], result.mape / 100, fmt="0.0%")
        _num(ws[f"F{i}"], result.rmse)
        _pct(ws[f"G{i}"], fc_cagr)

    # Freeze pane and column widths
    ws.freeze_panes = "A3"
    for col, width in zip("ABCDEFGH", [22, 14, 4, 28, 10, 14, 14, 4]):
        ws.column_dimensions[get_column_letter(ord(col) - 64)].width = width


def _forecast_sheet(ws, result: ForecastResult, historical: pd.Series):
    ws.title = result.model_name[:31]
    ws.sheet_view.showGridLines = False

    headers = ["Period", "Historical", "Fitted", "Forecast",
               "CI Lower (80%)", "CI Upper (80%)", "Residual %"]
    for c_idx, h in enumerate(headers, 1):
        _hdr(ws.cell(1, c_idx), h, bg=C_TITLE_BG)
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    # Historical + fitted rows
    for r_idx, (idx, val) in enumerate(historical.items(), 2):
        ws.cell(r_idx, 1).value = str(idx)[:20]
        ws.cell(r_idx, 1).font  = Font(size=9, name="Arial")
        _num(ws.cell(r_idx, 2), round(float(val), 2), color=C_INPUT_FG)

        fitted_val = result.fitted.get(idx)
        if fitted_val is not None and not np.isnan(fitted_val):
            _num(ws.cell(r_idx, 3), round(float(fitted_val), 2))
            if val != 0:
                resid = (val - fitted_val) / val
                _pct(ws.cell(r_idx, 7), round(float(resid), 4),
                     color=(C_BEAR if abs(resid) > 0.10 else C_FORMULA_FG))

        fill = PatternFill("solid", start_color=(C_ACCENT if r_idx % 2 == 0 else "FFFFFF"))
        for c in range(1, 8):
            ws.cell(r_idx, c).fill = fill

    # Forecast rows
    offset = len(historical) + 2
    for f_idx, (idx, val) in enumerate(result.forecast.items()):
        row = offset + f_idx
        ws.cell(row, 1).value = str(idx)[:20]
        ws.cell(row, 1).font  = Font(size=9, name="Arial")
        _num(ws.cell(row, 4), round(float(val), 2), color=C_LINK_FG)
        _num(ws.cell(row, 5), round(float(result.confidence_lo.iloc[f_idx]), 2))
        _num(ws.cell(row, 6), round(float(result.confidence_hi.iloc[f_idx]), 2))
        fill = PatternFill("solid", start_color=(C_BULL if f_idx % 2 == 0 else "F0FFF4"))
        for c in range(1, 7):
            ws.cell(row, c).fill = fill

    ws.freeze_panes = "A2"

    # Inline chart
    last_hist_row = len(historical) + 1
    last_fc_row   = last_hist_row + len(result.forecast)
    if last_fc_row > 2:
        chart = LineChart()
        chart.title  = f"{result.model_name} — Forecast"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Period"
        chart.width  = 20
        chart.height = 12

        hist_data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=last_hist_row)
        fc_data   = Reference(ws, min_col=4, max_col=4, min_row=1, max_row=last_fc_row)
        chart.add_data(hist_data, titles_from_data=True)
        chart.add_data(fc_data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "1F4E79"
        chart.series[0].graphicalProperties.line.width = 20000
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.line.solidFill = "2E75B6"
            chart.series[1].graphicalProperties.line.dashDot = "dash"

        ws.add_chart(chart, "I2")


def _scenario_sheet(ws, bundle: ScenarioBundle):
    ws.title = "Scenarios"
    ws.sheet_view.showGridLines = False

    headers = ["Period", "Historical", "Bear Case", "Base Case", "Bull Case",
               "Bear Growth", "Base Growth", "Bull Growth"]
    for c_idx, h in enumerate(headers, 1):
        _hdr(ws.cell(1, c_idx), h, bg=C_TITLE_BG)
        ws.column_dimensions[get_column_letter(c_idx)].width = 16

    # Historical
    for r_idx, (idx, val) in enumerate(bundle.historical.items(), 2):
        ws.cell(r_idx, 1).value = str(idx)[:20]
        _num(ws.cell(r_idx, 2), round(float(val), 2), color=C_INPUT_FG)
        fill = PatternFill("solid", start_color=(C_ACCENT if r_idx % 2 == 0 else "FFFFFF"))
        for c in range(1, 9): ws.cell(r_idx, c).fill = fill

    # Forecasts
    offset = len(bundle.historical) + 2
    for f_idx in range(len(bundle.base.forecast)):
        row = offset + f_idx
        idx = bundle.base.forecast.index[f_idx]
        ws.cell(row, 1).value = str(idx)[:20]

        bear_v = float(bundle.bear.forecast.iloc[f_idx])
        base_v = float(bundle.base.forecast.iloc[f_idx])
        bull_v = float(bundle.bull.forecast.iloc[f_idx])

        _num(ws.cell(row, 3), round(bear_v, 2), color=C_FORMULA_FG)
        _num(ws.cell(row, 4), round(base_v, 2), color=C_FORMULA_FG)
        _num(ws.cell(row, 5), round(bull_v, 2), color=C_FORMULA_FG)

        # Growth vs prior period
        if f_idx == 0:
            prior = float(bundle.historical.iloc[-1])
        else:
            prior_bear = float(bundle.bear.forecast.iloc[f_idx - 1])
            prior_base = float(bundle.base.forecast.iloc[f_idx - 1])
            prior_bull = float(bundle.bull.forecast.iloc[f_idx - 1])
            if prior_base != 0:
                _pct(ws.cell(row, 6), (bear_v / prior_bear - 1) if prior_bear else 0)
                _pct(ws.cell(row, 7), (base_v / prior_base - 1) if prior_base else 0)
                _pct(ws.cell(row, 8), (bull_v / prior_bull - 1) if prior_bull else 0)

        fill = PatternFill("solid", start_color=(C_BULL if f_idx % 2 == 0 else "F0FFF4"))
        for c in range(1, 9): ws.cell(row, c).fill = fill

    ws.freeze_panes = "A2"


# ------------------------------------------------------------------ #
#  Main report builder                                                 #
# ------------------------------------------------------------------ #
def export_report(series: pd.Series,
                  results: dict[str, ForecastResult],
                  bundle: ScenarioBundle,
                  output_path: str = "forecast_report.xlsx") -> str:
    wb = Workbook()

    # Summary
    _summary_sheet(wb.active, series, results, bundle, series.name or "Revenue")

    # One sheet per model
    for name, result in results.items():
        ws = wb.create_sheet()
        _forecast_sheet(ws, result, series)

    # Scenario sheet
    ws = wb.create_sheet()
    _scenario_sheet(ws, bundle)

    # Ensemble
    ensemble = EnsembleForecaster()
    ensemble.results = results
    ens_fc = ensemble.ensemble_forecast()
    if ens_fc is not None:
        ws_ens = wb.create_sheet("Ensemble_Forecast")
        _hdr(ws_ens["A1"], "Period",             bg=C_TITLE_BG)
        _hdr(ws_ens["B1"], "Ensemble Forecast",  bg=C_TITLE_BG)
        for c_idx in [1, 2]:
            ws_ens.column_dimensions[get_column_letter(c_idx)].width = 20
        for r_idx, (idx, val) in enumerate(ens_fc.items(), 2):
            ws_ens.cell(r_idx, 1).value = str(idx)[:20]
            _num(ws_ens.cell(r_idx, 2), round(float(val), 2), color=C_LINK_FG)

    wb.save(output_path)
    return output_path
